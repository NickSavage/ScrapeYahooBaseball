#!/usr/bin/env python

import mechanize
import cookielib
from bs4 import BeautifulSoup
import csv
import datetime
import sys
import configobj
import argparse
import openpyxl

CONFIG_LEAGUEID = 0
CONFIG_FILENAME = ""
CONFIG_SORT = ""
CONFIG_CSV = True
CONFIG_EXCEL = False
CONFIG_BOTH = False
writer = None
ofile = None
workbook = None
worksheet = None
currentRow = 1
excelFilename = None

def main():
        global CONFIG_LEAGUEID
        global CONFIG_FILENAME
        global CONFIG_SORT 
        global CONFIG_EXCEL
        global CONFIG_CSV
        global CONFIG_BOTH
        # command line parsing
        parser = argparse.ArgumentParser(description="Scrape baseball player data from Yahoo Fantasy Sports")
        parser.add_argument('-c','--config', default="config.ini")
        parser.add_argument('-m','--max-pages', type=int, default=1)
        parser.add_argument('-t','--timeframe', choices=['2016', '2015', 'today', '7', '14', '30'], default='2016')
        parser.add_argument('-s','--sort', choices=['OR','AR'], default="AR")
        parser.add_argument('-p','--password')
        parser.add_argument('-u','--username')
        parser.add_argument('-l','--leagueid', type=int)
        parser.add_argument('--available', action='store_true')
        parser.add_argument('--excel', action='store_true')
        parser.add_argument('action', choices=['pitchers', 'batters', 'both'], help="scrape pitchers")
        
        args = parser.parse_args()
        action = str(args.action)

        if (action == "pitchers"):
                CONFIG_PLAYERTYPE = "1"
        elif (action == "batters"):
                CONFIG_PLAYERTYPE = "2"
        elif (action == "both"):
                CONFIG_PLAYERTYPE = "3"
        CONFIG_FILENAME = args.config
        maxPages = args.max_pages
        CONFIG_TIMEFRAME = args.timeframe
        CONFIG_SORT = args.sort
        if (args.available):
                CONFIG_AVAILABLE = 2
        elif (args.available == False):
                CONFIG_AVAILABLE = 1

        if (args.excel == True):
                CONFIG_EXCEL = True
                CONFIG_CSV = False
     # load config from config file. At the moment, its only for leagueID, username and password
        # but in the future it will include overriding the defaults of the options
        (CONFIG_LEAGUEID, username, password) = loadConfig(CONFIG_FILENAME)
        username = args.username
        password = args.password
        CONFIG_LEAGUEID = args.leagueid
        if (username == None) or (password == None):
                raise Exception("No username or password entered. Exiting...")
        if (CONFIG_LEAGUEID == None):
                raise Exception("No League ID entered. Exiting...")        
        # Start scraping process
	br = authentication(username, password)
        if CONFIG_PLAYERTYPE == "1": 
                scrape(br, "1", CONFIG_TIMEFRAME, CONFIG_AVAILABLE, maxPages)
        elif CONFIG_PLAYERTYPE == "2":
                scrape(br, "2", CONFIG_TIMEFRAME, CONFIG_AVAILABLE, maxPages)
        elif CONFIG_PLAYERTYPE == "3": 
                scrape(br, "1", CONFIG_TIMEFRAME, CONFIG_AVAILABLE, maxPages)
                CONFIG_BOTH = True
                scrape(br, "2", CONFIG_TIMEFRAME, CONFIG_AVAILABLE, maxPages)
                
def loadConfig(filename):
        try:
                filename = open(filename, 'r')
        except:
                print "No configuration file."
                return(0, "", "")
        config = configobj.ConfigObj(filename)

        if (config['username']):
                username = config['username']
        if (config['password']):
                password = config['password']
        if (config['leagueID']):
                leagueID = config['leagueID']

        return(leagueID, username, password)
 
def buildWriter(CONFIG_PLAYERTYPE):
        global CONFIG_LEAGUEID
        global CONFIG_CSV
        global CONFIG_EXCEL
        global writer
        global ofile
        global workbook
        global worksheet
        global excelFilename
        global CONFIG_BOTH
        global currentRow
        if (CONFIG_CSV == True):
                filename = buildFilename("csv", CONFIG_PLAYERTYPE)
                ofile = open(filename, "wb")
                writer = csv.writer(ofile, delimiter = ',', escapechar = ' ')
        if (CONFIG_EXCEL == True):
                if (workbook != None):
                        currentRow = 1
                        worksheet = workbook.create_sheet()
                if (workbook == None):
                        workbook = openpyxl.Workbook()
                        worksheet = workbook.active
                        excelFilename = buildFilename("excel", CONFIG_PLAYERTYPE)
                print excelFilename
                if CONFIG_PLAYERTYPE == "1":
                        worksheet.title = 'Pitchers'
                if CONFIG_PLAYERTYPE == "2":
                        worksheet.title = 'Batters'
def addWorksheet(CONFIG_PLAYERTYPE):
        global workbook
        global worksheet
        if CONFIG_PLAYERTYPE == "1":
                worksheet.title = 'Pitchers'
        if CONFIG_PLAYERTYPE == "2":
                worksheet.title = 'Batters'

        
def closeFile():
        global ofile
        global workbook
        global excelFilename
        global CONFIG_CSV
        global CONFIG_EXCEL
        
        if (CONFIG_CSV == True):
                ofile.close()
        if (CONFIG_EXCEL == True):
                workbook.save(excelFilename)
                
def scrape(br, localPlayerType, localTimeFrame, localAvailable, localMaxPages):
        #building url for scraping

        buildWriter(localPlayerType)
        url = buildURL(localPlayerType, localTimeFrame, localAvailable)
	content = br.open(url + '0')
	soup = BeautifulSoup(content, "lxml")
        print url

        # build list of stats to scrape, based on what your league counts
	statsList = soup.findAll('th', {'class':'Ta-end'})
	stats = ['Name', 'Team', 'Pos', 'Fantasy Team']
	for s in statsList:
		t = str(s.findAll(text=True))
		t = t[3:len(t)-2]
		stats.append(t.split(",")[0])
		try:
                       stats.remove("Rankings")
		except:
			continue
	writeData(stats) # write first row of csv file
	pageNum = 0 # initialize counter for scraping lists of players

	while True:
		count = 0
		pageCount = str(pageNum * 25)
                print "Loading page",(pageNum+1)

                end = scrapePage(count, pageCount, br, url, stats)
                count += 1
                pageNum += 1
                if pageNum >= localMaxPages: break
                if end == 1: break
        closeFile()
        # writeData exists only for the future where options will exist other than
        # just CSV files, mainly excel.
def writeData(data):
        global writer
        global worksheet
        global currentRow
        global CONFIG_CSV
        global CONFIG_EXCEL

        if (CONFIG_CSV == True):
                writer.writerow(data)
        if (CONFIG_EXCEL == True):
                for i in range(1,len(data)):
                        cell = worksheet.cell(row = currentRow, column = i)
                        cell.value = data[i-1]
                currentRow += 1
       
def authentication(username, password):
	cj = cookielib.CookieJar()
	br = mechanize.Browser()
	br.set_handle_robots(False)
	br.addheaders = [('User-agent', 'Mozilla/5.0 (Windows; U; Windows NT 6.0; en-US; rv:1.9.0.6')]
	br.set_cookiejar(cj)
	br.open("https://login.yahoo.com/config/login_verify2?&.src=ym&.intl=us")
	br.select_form(nr=0)
	br.form["username"] = username
	submit_response = br.submit()
	br.select_form(nr=0)
	br.form["passwd"] = password
	br.submit()
        return br

def scrapePage(count, pageCount, br, url, stats):
        global writer
        # gets information from url
        content = br.open(url + pageCount)
        soup = BeautifulSoup(content, "lxml")
        
        players = soup.findAll('div', {'class':'ysf-player-name Nowrap Grid-u Relative Lh-xs Ta-start'})
        dataList = soup.findAll('td', {'class': 'Ta-end'})

        # Following block finds all the fantasy team names and puts them into fanTeams without all the html and formatting
        fantasyTeams = soup.findAll('div', {'style':'text-overflow: ellipsis; overflow: hidden;'})
        # Make array of fantasy team names
        fanTeams = []
        for f in fantasyTeams:
                s = f.findAll(text=True)
                fanTeams.append(str(s)[3:-2])
        fanTeams.pop(0)

        # Find end of player list
        try:
                str(players[0].findAll(text=True))
        except:
                print "End of Players"
                return 1

        for player in players:
                playerStats = []
                # extracts only name, team, position from html.
                # magic python function!
                playerData = str(player.findAll(text=True)) 
#                name = getName(playerData)
                (name, team, pos) = getNameTeamAndPosition(playerData)

                fanTeam = fanTeams[0]
                fanTeams.pop(0)
                playerStats.extend([name, team, pos, fanTeam])
                for i in range(0, len(stats)-4):
                        tmp = str(dataList[0].findAll(text=True))
                        playerStats.append(tmp[3:-2])
                        dataList.pop(0)
                writeData(playerStats)
        
def buildURL(type, time, available):
        global CONFIG_LEAGUEID
        global CONFIG_SORT
	begin_url = 'http://baseball.fantasysports.yahoo.com/b1/' + str(CONFIG_LEAGUEID) + '/players?status='
	end_url = '&myteam=0&sort=' + str(CONFIG_SORT) + '&sdir=1&count='

	if available == 1: status = 'ALL'
	if available == 2: status = 'A'
	if type == "1": pos = 'P'
	if type == "2": pos = 'B'
	if time == '2016': timeFrame = 'S_2016'
	if time == '2015': timeFrame = 'S_2015'
	if time == '30': timeFrame = 'L30'
	if time == '14': timeFrame = 'L14'
	if time == '7': timeFrame = 'L7'
	if time == 'today': timeFrame = 'L'

	mid_url = status + '&pos=' + pos + '&cut_type=33&stat1=S_' + timeFrame
	return begin_url + mid_url + end_url

def getNameTeamAndPosition(data):
	playerData = data.split("'")
	if data[2] == '"':
		teampos = playerData[4]
	else:
		teampos = playerData[5]
        name = fixText(playerData[1])
	team = teampos[0:teampos.find("-")-1]
	pos = teampos[teampos.find("-")+2:len(teampos)]
	return (name, team, pos)

def buildFilename(dataType, CONFIG_PLAYERTYPE):
        global CONFIG_LEAGUEID
        global excelFilename
        global CONFIG_CSV
        global CONFIG_EXCEL
        # build filename for export
        filename = ""
        if (CONFIG_CSV == True):
                if CONFIG_PLAYERTYPE == "1":
                        end_filename = 'Pitchers'
                if CONFIG_PLAYERTYPE == "2":
                        end_filename = 'Batters'
                        
                filename = 'FBB_data_' + str(CONFIG_LEAGUEID) + '_' + end_filename + '_' + str(datetime.date.today()) + '.csv'
        elif (CONFIG_EXCEL == True):
                filename = 'FBB_data_' + str(CONFIG_LEAGUEID) + '_' + str(datetime.date.today()) + '.xlsx'
        return filename

def fixText(str):
	s = str
	s = s.replace('\\xe1', 'a')
	s = s.replace('\\xe0', 'a')
	s = s.replace('\\xc1', 'A')
	s = s.replace('\\xe9', 'e')
	s = s.replace('\\xc9', 'E')
	s = s.replace('\\xed', 'i')
	s = s.replace('\\xcd', 'I')
	s = s.replace('\\xf3', 'o')
	s = s.replace('\\xd3', 'O')
	s = s.replace('\\xfa', 'u')
	s = s.replace('\\xda', 'U')
	s = s.replace('\\xf1', 'n')
	return s

if __name__ == "__main__":
	main()
